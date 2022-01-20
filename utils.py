import openpyxl


def getrowcount(file,sheetName):
         workbook=openpyxl.load_workbook(file)
         sheet=workbook.get_sheet_by_name(sheetName)
         return sheet.max_row

def getrolcount(file, sheetName):
        workbook = openpyxl.load_workbook(file)
        sheet = workbook.get_sheet_by_name(sheetName)
        return sheet.max_col

def readdata(file, sheetName,rownum,colnum):
        workbook = openpyxl.load_workbook(file)
        sheet = workbook.get_sheet_by_name(sheetName)
        return sheet.cell(row=rownum,column=colnum).value


def writedata(file, sheetName, rownum, colnum, data):
        workbook = openpyxl.load_workbook(file)
        sheet = workbook.get_sheet_by_name(sheetName)
        return sheet.cell(row=rownum, column=colnum).value == data
        workbook.save(file)
